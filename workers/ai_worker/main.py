from __future__ import annotations

import os
from functools import lru_cache
from pathlib import Path
from typing import List

from fastapi import FastAPI
from openpyxl import load_workbook
from pydantic import BaseModel
from pypdf import PdfReader
from sentence_transformers import SentenceTransformer
from docx import Document as WordDocument

app = FastAPI(title="AiCan Worker", version="0.1.0")


class ExtractRequest(BaseModel):
    file_path: str


class ExtractResponse(BaseModel):
    file_path: str
    extracted_text: str


class ClassifyRequest(BaseModel):
    file_name: str
    extracted_text: str


class ClassifyResponse(BaseModel):
    category: str
    confidence: float
    customer_name: str


class EmbedRequest(BaseModel):
    text: str
    input_type: str = "passage"
    model: str | None = None


class EmbedBatchRequest(BaseModel):
    texts: List[str]
    input_type: str = "passage"
    model: str | None = None


class EmbedResponse(BaseModel):
    embedding: List[float]


class EmbedBatchResponse(BaseModel):
    embeddings: List[List[float]]


def classify(file_name: str, extracted_text: str) -> ClassifyResponse:
    corpus = f"{file_name} {extracted_text}".lower()
    if "invoice" in corpus:
        return ClassifyResponse(category="invoice", confidence=0.92, customer_name="unassigned")
    if "hr" in corpus or "employee" in corpus:
        return ClassifyResponse(category="hr", confidence=0.88, customer_name="internal")
    return ClassifyResponse(category="general", confidence=0.71, customer_name="unassigned")


@lru_cache(maxsize=2)
def load_embedder(model_name: str) -> SentenceTransformer:
    return SentenceTransformer(model_name, device=os.getenv("AICAN_EMBEDDING_DEVICE", "cpu"))


def prefix_texts(texts: List[str], input_type: str) -> List[str]:
    prefix = "query: " if input_type.lower() == "query" else "passage: "
    return [text if text.startswith(prefix) else f"{prefix}{text}" for text in texts]


def embed_texts(texts: List[str], input_type: str, model_name: str | None) -> List[List[float]]:
    effective_model = model_name or os.getenv("AICAN_EMBEDDING_MODEL", "intfloat/multilingual-e5-base")
    embedder = load_embedder(effective_model)
    embeddings = embedder.encode(
        prefix_texts(texts, input_type),
        normalize_embeddings=True,
        convert_to_numpy=True,
        show_progress_bar=False,
    )
    return [row.tolist() for row in embeddings]


@app.on_event("startup")
async def warm_embedding_model() -> None:
    model_name = os.getenv("AICAN_EMBEDDING_MODEL", "intfloat/multilingual-e5-base")
    print(f"[AiCan Worker] Loading embedding model '{model_name}'...", flush=True)
    load_embedder(model_name)
    print(f"[AiCan Worker] Embedding model '{model_name}' is ready.", flush=True)


@app.get("/healthz")
async def healthz() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/extract", response_model=ExtractResponse)
async def extract(request: ExtractRequest) -> ExtractResponse:
    path = Path(request.file_path)
    if not path.exists():
        text = f"placeholder extracted text from {path.name}"
    else:
        extension = path.suffix.lower()
        if extension in {".txt", ".md", ".csv", ".json"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
        elif extension == ".pdf":
            text = extract_pdf_text(path)
        elif extension == ".docx":
            text = extract_docx_text(path)
        elif extension == ".xlsx":
            text = extract_xlsx_text(path)
        else:
            text = f"placeholder extracted text from {path.name}"
    return ExtractResponse(file_path=request.file_path, extracted_text=text)


@app.post("/classify", response_model=ClassifyResponse)
async def classify_endpoint(request: ClassifyRequest) -> ClassifyResponse:
    return classify(request.file_name, request.extracted_text)


@app.post("/embed", response_model=EmbedResponse)
async def embed(request: EmbedRequest) -> EmbedResponse:
    embedding = embed_texts([request.text], request.input_type, request.model)[0]
    return EmbedResponse(embedding=embedding)


@app.post("/embed-batch", response_model=EmbedBatchResponse)
async def embed_batch(request: EmbedBatchRequest) -> EmbedBatchResponse:
    embeddings = embed_texts(request.texts, request.input_type, request.model)
    return EmbedBatchResponse(embeddings=embeddings)


def extract_pdf_text(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n\n".join(page.strip() for page in pages if page.strip())
        return text or f"placeholder extracted text from {path.name}"
    except Exception:
        return f"placeholder extracted text from {path.name}"


def extract_docx_text(path: Path) -> str:
    try:
        document = WordDocument(str(path))
        lines = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
        text = "\n".join(lines)
        return text or f"placeholder extracted text from {path.name}"
    except Exception:
        return f"placeholder extracted text from {path.name}"


def extract_xlsx_text(path: Path) -> str:
    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        parts: List[str] = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                parts.append(f"[Sheet: {sheet.title}]\n" + "\n".join(rows))
        workbook.close()
        text = "\n\n".join(parts)
        return text or f"placeholder extracted text from {path.name}"
    except Exception:
        return f"placeholder extracted text from {path.name}"
